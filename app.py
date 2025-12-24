import streamlit as st
import pandas as pd
import io
import os
import google.generativeai as genai
import json
import zipfile
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
st.set_page_config(
    page_title="HostAfrica AI Auditor - Enhanced Security Edition V2", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- CUSTOM STYLES ---
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
    }
    .strength-item {
        background-color: #d4edda;
        padding: 0.5rem;
        margin: 0.3rem 0;
        border-radius: 0.3rem;
        border-left: 3px solid #28a745;
    }
    .improvement-item {
        background-color: #fff3cd;
        padding: 0.5rem;
        margin: 0.3rem 0;
        border-radius: 0.3rem;
        border-left: 3px solid #ffc107;
    }
    .critical-item {
        background-color: #f8d7da;
        padding: 0.5rem;
        margin: 0.3rem 0;
        border-radius: 0.3rem;
        border-left: 3px solid #dc3545;
    }
</style>
""", unsafe_allow_html=True)

# --- GEMINI AI SETUP ---
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY", "")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
    # Using gemini-2.5-flash-lite for better free tier availability
    model = genai.GenerativeModel('gemini-2.5-flash-lite')

# --- DATA STRUCTURES ---
if 'agents' not in st.session_state:
    st.session_state.agents = {}

def get_initial_agent(name=""):
    return {
        "name": name,
        "audit_data": None,
        "total_chats": 0,
        "audit_timestamp": None,
        "raw_transcripts": [],
        "review_period_days": REVIEW_PERIOD_DAYS,
        "chat_dates": []
    }

# --- RECURSIVE ZIP PROCESSING ---
def get_agent_transcripts(uploaded_zip, target_name):
    """Extract transcripts for a specific agent from ZIP file"""
    transcripts = []
    chat_metadata = []
    
    with zipfile.ZipFile(uploaded_zip, 'r') as z:
        for file_path in z.namelist():
            if file_path.endswith('.json'):
                with z.open(file_path) as f:
                    try:
                        data = json.load(f)
                        chat_text = ""
                        belongs_to_agent = False
                        message_count = 0
                        
                        # Extract metadata
                        chat_id = data.get("id", "unknown")
                        started_at = data.get("started", "")
                        
                        # Extracting transcript data based on tawk.to JSON structure
                        for msg in data.get("messages", []):
                            name = msg.get("sender", {}).get("n", "Visitor")
                            body = msg.get("msg", "")
                            timestamp = msg.get("t", "")
                            
                            chat_text += f"[{timestamp}] {name}: {body}\n"
                            message_count += 1
                            
                            if name == target_name:
                                belongs_to_agent = True
                        
                        if belongs_to_agent and message_count > 3:  # Only include substantial chats
                            transcripts.append(chat_text)
                            chat_metadata.append({
                                "chat_id": chat_id,
                                "started_at": started_at,
                                "message_count": message_count
                            })
                    except Exception as e:
                        st.warning(f"Could not process file {file_path}: {str(e)}")
                        continue
    
    return transcripts, chat_metadata

# --- ENHANCED AI AUDIT LOGIC ---
def run_comprehensive_audit(transcripts, agent_name, total_chats, review_period_days):
    """Run comprehensive AI-powered audit with detailed analysis"""
    
    # Calculate chat volume metrics
    
    # Use up to 50 transcripts for comprehensive analysis
    sample_size = min(50, len(transcripts))
    sample = "\n\n========== NEW CHAT SESSION ==========\n\n".join(transcripts[:sample_size])
    
    prompt = f"""
You are a Senior Technical QA Auditor at HostAfrica with 10+ years of experience evaluating technical support quality.
You are conducting a comprehensive performance review of agent: {agent_name}

REVIEW CONTEXT:
- Total chats analyzed: {total_chats}
- Review period: {review_period_days} days (quarterly review)

CRITICAL INSTRUCTIONS:
1. Analyze ONLY the agent's performance, NOT bots or automated messages
2. Focus on REAL examples from the actual chat transcripts provided
3. All scores must use the correct scale as specified
4. Provide actionable, specific feedback based on actual chat evidence
5. Pay special attention to PIN verification protocol - this is a critical security measure

SCORING SCALES:
- Individual metrics: 0.0 to 5.0 (where 5.0 is exceptional)
- Overall score: 0.0 to 10.0 (composite of all metrics)

EVALUATION FRAMEWORK (Weighted):

1. SECURITY & PIN VERIFICATION PROTOCOL (Weight: 20%)
   CRITICAL FOCUS AREAS:
   - Did agent request PIN when initiating new chats OR when taking over from bot?
   - Did agent avoid redundant PIN requests (not asking for already-provided PIN in SAME session)?
   - Did agent follow proper security procedures consistently?
   - IMPORTANT: In transferred chats, agent MUST re-verify PIN in their own session before account actions
   - Flag any instances where account modifications were made without explicit PIN verification
   Rate: 0.0-5.0

2. TECHNICAL CAPABILITY & ACCURACY (Weight: 25%)
   - Accuracy in diagnosing DNS, Email, SSL, WordPress, hosting issues
   - Proper use of diagnostic tools (Ping, Traceroute, WHOIS, cPanel)
   - Correctness of technical solutions provided
   - Depth of technical knowledge demonstrated
   Rate: 0.0-5.0

3. COMMUNICATION & PROFESSIONALISM (Weight: 15%)
   - Clarity and professionalism in communication
   - Empathy and patience with customers (especially frustrated ones)
   - Grammar, spelling, and tone appropriateness
   - De-escalation techniques for difficult situations
   Rate: 0.0-5.0

4. INVESTIGATIVE & PROBLEM-SOLVING APPROACH (Weight: 20%)
   - Systematic troubleshooting methodology
   - Asking relevant diagnostic questions
   - Root cause analysis capability
   - Thoroughness in investigation
   - Proactive information gathering
   Rate: 0.0-5.0

5. CHAT OWNERSHIP & RESOLUTION (Weight: 20%)
   - Taking full ownership of issues
   - Following through to resolution
   - Proactive communication and updates
   - Proper escalation when needed
   - Ensuring customer satisfaction
   Rate: 0.0-5.0

OUTPUT REQUIREMENTS:
Provide exactly 20 detailed technical examples from the actual transcripts. Each example must:
- Reference a REAL issue from the chats
- Show the ACTUAL agent action/response
- Include specific improvement recommendations
- Indicate PIN handling quality (Yes/No/Redundant/N/A)
- Classify severity (Minor/Moderate/Major/Critical)

IMPORTANT: For PIN verification:
- "Yes" = Agent properly requested/verified PIN before account actions
- "No" = Agent performed account actions WITHOUT proper PIN verification (SECURITY RISK)
- "Redundant" = Agent asked for PIN that was already provided in same session
- "N/A" = No account access required (e.g., general questions, pre-sales)

Return ONLY valid JSON in this exact structure:
{{
    "overall_score": 0.0,
    "overall_assessment": "Comprehensive 3-4 paragraph summary of agent's performance, highlighting key patterns observed across all interactions. Include commentary on chat volume performance.",
    "metrics": {{
        "security_pin_protocol": 0.0,
        "technical_capability": 0.0,
        "communication_professionalism": 0.0,
        "investigative_approach": 0.0,
        "chat_ownership_resolution": 0.0
    }},
    "key_strengths": [
        "Specific strength with example from chats",
        "Specific strength with example from chats",
        "Specific strength with example from chats",
        "Specific strength with example from chats",
        "Specific strength with example from chats"
    ],
    "key_development_areas": [
        "Specific area for improvement with actionable advice",
        "Specific area for improvement with actionable advice",
        "Specific area for improvement with actionable advice",
        "Specific area for improvement with actionable advice",
        "Specific area for improvement with actionable advice"
    ],
    "pin_protocol_feedback": "DETAILED analysis of PIN verification practices across all chats. Specifically note: (1) How many chats had proper PIN verification, (2) How many had PIN bypasses, (3) Pattern analysis - does agent verify in new chats but skip in transferred chats?, (4) Specific examples of good and poor PIN handling, (5) Security risk assessment",
    "technical_examples": [
        {{
            "example_number": 1,
            "client_name": "Customer's name from the chat (e.g., 'John Doe', 'Sarah Smith')",
            "pin_number": "PIN number if mentioned in chat (e.g., '1234', 'Not provided', 'N/A')",
            "issue_type": "VPS/Server, Domain/WHOIS, DNS, Email, SSL, WordPress, cPanel, Billing, etc.",
            "customer_issue": "Specific detailed customer complaint or issue",
            "agent_action": "Detailed description of what agent actually did or said",
            "pin_handled_well": "Yes/No/Redundant/N/A",
            "outcome": "What happened as a result of agent's action",
            "assessment": "Critical evaluation - was this handled well or poorly? Why?",
            "improvement": "Specific actionable improvement suggestion or 'None' if handled perfectly",
            "severity": "Minor/Moderate/Major/Critical"
        }},
        {{
            "example_number": 2,
            ... (continue for 20 examples total)
        }}
    ],
    "performance_trends": {{
        "response_time_assessment": "Analysis of agent's response speed and communication timing",
        "consistency": "How consistent is the agent's performance across different issue types",
        "technical_depth": "Assessment of technical knowledge depth and problem-solving capability",
        "customer_satisfaction_indicators": "Signs of customer satisfaction or frustration based on chat outcomes"
    }},
    "recommended_training": [
        "Specific training recommendation based on identified gaps (minimum 3)",
        "Specific training recommendation based on identified gaps",
        "Specific training recommendation based on identified gaps"
    ],
    "standout_moments": [
        "Exceptional handling example with specific details from chats",
        "Exceptional handling example with specific details from chats"
    ],
    "critical_incidents": [
        "Any critical errors, serious security lapses, or major issues (be specific)"
    ],
}}

CHAT TRANSCRIPTS TO ANALYZE:
{sample}

Remember: Base ALL examples and assessments on the ACTUAL transcripts provided above. Be specific, fair, and constructive. Focus heavily on PIN verification protocol as this is a critical security concern for HostAfrica.
"""
    
    try:
        response = model.generate_content(prompt)
        text = response.text
        
        # Clean up JSON response
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            text = text.split("```")[1].split("```")[0]
        
        # Parse JSON
        audit_result = json.loads(text.strip())
        
        # Validate and cap scores
        audit_result['overall_score'] = min(float(audit_result.get('overall_score', 0)), 10.0)
        for key in audit_result.get('metrics', {}):
            audit_result['metrics'][key] = min(float(audit_result['metrics'][key]), 5.0)
        
        # Add volume metrics to audit result
        
        return audit_result
        
    except json.JSONDecodeError as e:
        st.error(f"JSON Parsing Error: {e}")
        st.error(f"Raw response: {text[:500]}")
        return None
    except Exception as e:
        st.error(f"AI Generation Error: {e}")
        return None

# --- ENHANCED PDF REPORT GENERATION ---
def generate_pdf_report(agent_data, agent_name, output_path):
    """Generate a comprehensive PDF performance review report"""
    
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=24,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading1_style = ParagraphStyle(
        'CustomHeading1',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    heading2_style = ParagraphStyle(
        'CustomHeading2',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        spaceAfter=10,
        spaceBefore=10
    )
    
    body_style = ParagraphStyle(
        'CustomBody',
        parent=styles['BodyText'],
        fontSize=11,
        alignment=TA_JUSTIFY,
        spaceAfter=10
    )
    
    # Build story
    story = []
    
    # Get volume metrics
    
    # Title
    story.append(Paragraph("PERFORMANCE REVIEW REPORT", title_style))
    story.append(Paragraph(f"<b>Agent:</b> {agent_name}", styles['Normal']))
    story.append(Paragraph(f"<b>Review Date:</b> {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
    story.append(Paragraph(f"<b>Chats Analyzed:</b> {agent_data.get('total_chats', 0)}", styles['Normal']))
    story.append(Paragraph(f"<b>Review Period:</b> {agent_data.get('review_period_days', 90)} days (Quarterly)", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Chat Volume Performance Section
        
        volume_data = [
            ['Metric', 'Value', 'Assessment'],
        ]
        
        volume_table = Table(volume_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        
        # Color code based on performance
        rating_color = colors.HexColor('#28a745')  # Green
            rating_color = colors.HexColor('#ffc107')  # Yellow
            rating_color = colors.HexColor('#dc3545')  # Red
        
        volume_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
            ('BACKGROUND', (2, 3), (2, 3), rating_color),
            ('TEXTCOLOR', (2, 3), (2, 3), colors.white)
        ]))
        story.append(volume_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Overall Score
    overall_score = agent_data.get('audit_data', {}).get('overall_score', 0)
    story.append(Paragraph("OVERALL PERFORMANCE SCORE", heading1_style))
    
    score_data = [[f"{overall_score}/10.0"]]
    score_table = Table(score_data, colWidths=[2*inch])
    score_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 24),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#e8f4f8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1f77b4')),
        ('PADDING', (0, 0), (-1, -1), 20),
        ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1f77b4'))
    ]))
    story.append(score_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Overall Assessment
    story.append(Paragraph("OVERALL ASSESSMENT", heading1_style))
    assessment = agent_data.get('audit_data', {}).get('overall_assessment', 'No assessment available')
    story.append(Paragraph(assessment, body_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Metrics
    story.append(Paragraph("PERFORMANCE METRICS", heading1_style))
    metrics = agent_data.get('audit_data', {}).get('metrics', {})
    
    metrics_data = [['Metric', 'Score (out of 5.0)']]
    for key, value in metrics.items():
        metric_name = key.replace('_', ' ').title()
        metrics_data.append([metric_name, f"{value}/5.0"])
    
    metrics_table = Table(metrics_data, colWidths=[4*inch, 1.5*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Page Break
    story.append(PageBreak())
    
    # Strengths
    story.append(Paragraph("KEY STRENGTHS", heading1_style))
    strengths = agent_data.get('audit_data', {}).get('key_strengths', [])
    for i, strength in enumerate(strengths, 1):
        story.append(Paragraph(f"<b>{i}.</b> {strength}", body_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Development Areas
    story.append(Paragraph("AREAS FOR DEVELOPMENT", heading1_style))
    dev_areas = agent_data.get('audit_data', {}).get('key_development_areas', [])
    for i, area in enumerate(dev_areas, 1):
        story.append(Paragraph(f"<b>{i}.</b> {area}", body_style))
    story.append(Spacer(1, 0.2*inch))
    
    # PIN Protocol Feedback
    story.append(Paragraph("SECURITY & PIN PROTOCOL ANALYSIS", heading1_style))
    pin_feedback = agent_data.get('audit_data', {}).get('pin_protocol_feedback', 'No feedback available')
    story.append(Paragraph(pin_feedback, body_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Page Break
    story.append(PageBreak())
    
    # Technical Examples
    story.append(Paragraph("DETAILED TECHNICAL EXAMPLES", heading1_style))
    examples = agent_data.get('audit_data', {}).get('technical_examples', [])
    
    for i, example in enumerate(examples, 1):
        story.append(Paragraph(f"<b>Example {i}: {example.get('issue_type', 'N/A')}</b>", heading2_style))
        
        example_data = [
            ['Client Name:', example.get('client_name', 'N/A')],
            ['PIN Number:', example.get('pin_number', 'N/A')],
            ['Customer Issue:', example.get('customer_issue', 'N/A')],
            ['Agent Action:', example.get('agent_action', 'N/A')],
            ['PIN Handled Well:', example.get('pin_handled_well', 'N/A')],
            ['Outcome:', example.get('outcome', 'N/A')],
            ['Assessment:', example.get('assessment', 'N/A')],
            ['Improvement:', example.get('improvement', 'N/A')],
            ['Severity:', example.get('severity', 'N/A')]
        ]
        
        example_table = Table(example_data, colWidths=[1.5*inch, 5*inch])
        example_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('PADDING', (0, 0), (-1, -1), 8)
        ]))
        
        story.append(example_table)
        story.append(Spacer(1, 0.2*inch))
        
        # Page break every 4 examples
        if i % 4 == 0 and i < len(examples):
            story.append(PageBreak())
    
    # Performance Trends (if available)
    trends = agent_data.get('audit_data', {}).get('performance_trends', {})
    if trends:
        story.append(PageBreak())
        story.append(Paragraph("PERFORMANCE TRENDS", heading1_style))
        for key, value in trends.items():
            trend_name = key.replace('_', ' ').title()
            story.append(Paragraph(f"<b>{trend_name}:</b> {value}", body_style))
        story.append(Spacer(1, 0.2*inch))
    
    # Recommended Training
    training = agent_data.get('audit_data', {}).get('recommended_training', [])
    if training:
        story.append(Paragraph("RECOMMENDED TRAINING", heading1_style))
        for i, recommendation in enumerate(training, 1):
            story.append(Paragraph(f"<b>{i}.</b> {recommendation}", body_style))
        story.append(Spacer(1, 0.2*inch))
    
    # Standout Moments
    standout = agent_data.get('audit_data', {}).get('standout_moments', [])
    if standout:
        story.append(Paragraph("STANDOUT MOMENTS", heading1_style))
        for i, moment in enumerate(standout, 1):
            story.append(Paragraph(f"<b>{i}.</b> {moment}", body_style))
        story.append(Spacer(1, 0.2*inch))
    
    # Critical Incidents
    critical = agent_data.get('audit_data', {}).get('critical_incidents', [])
    if critical:
        story.append(Paragraph("CRITICAL INCIDENTS", heading1_style))
        for i, incident in enumerate(critical, 1):
            story.append(Paragraph(f"<b>{i}.</b> {incident}", body_style))
    
    # Build PDF
    doc.build(story)
    return output_path

# --- EXCEL REPORT GENERATION (Consolidated Single-Sheet Version) ---
def generate_excel_report(agent_data, agent_name, output_path):
    """Generate a single-sheet Excel performance review report for easy copy-pasting"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Review Report"
    
    # Define styles
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16, color="1F77B4")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    current_row = 1

    # --- SECTION 1: HEADER & AGENT INFO ---
    ws.cell(row=current_row, column=1, value="PERFORMANCE REVIEW REPORT").font = title_font
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.cell(row=current_row, column=1).alignment = center_align
    current_row += 2
    
    info_items = [
        ("Agent:", agent_name),
        ("Review Date:", datetime.now().strftime('%B %d, %Y')),
        ("Chats Analyzed:", agent_data.get('total_chats', 0)),
        ("Review Period:", f"{agent_data.get('review_period_days', 90)} days (Quarterly)")
    ]
    
    for label, value in info_items:
        ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    current_row += 1

    # --- SECTION 2: OVERALL SCORE ---
    ws.cell(row=current_row, column=1, value="OVERALL PERFORMANCE SCORE").font = subheader_font
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 1
    
    overall_score = agent_data.get('audit_data', {}).get('overall_score', 0)
    score_cell = ws.cell(row=current_row, column=1, value=f"{overall_score}/10.0")
    score_cell.font = Font(bold=True, size=24, color="1F77B4")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=4)
    score_cell.alignment = center_align
    current_row += 3

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        current_row += 1
        
        # Note: I'm using the global EXPECTED_CHATS_PER_5_DAYS if available
        try:
            target_str = f"Target: {EXPECTED_CHATS_PER_5_DAYS}/week"
        except NameError:
            target_str = "Target: Standard/week"

        volume_data = [
            ['Metric', 'Value', 'Assessment'],
        ]
        
        for r_idx, row_data in enumerate(volume_data):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row + r_idx, column=c_idx, value=value)
                cell.border = border
                if r_idx == 0:
                    cell.fill, cell.font, cell.alignment = header_fill, header_font, center_align
                else:
                    cell.alignment = center_align if c_idx > 1 else left_align
        current_row += len(volume_data) + 1

    # --- SECTION 4: PERFORMANCE METRICS ---
    metrics = agent_data.get('audit_data', {}).get('metrics', {})
    if metrics:
        ws.cell(row=current_row, column=1, value="PERFORMANCE METRICS").font = subheader_font
        current_row += 1
        
        headers = ["Metric", "Score (out of 5.0)"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        
        current_row += 1
        for key, value in metrics.items():
            ws.cell(row=current_row, column=1, value=key.replace('_', ' ').title()).border = border
            score_cell = ws.cell(row=current_row, column=2, value=f"{value}/5.0")
            score_cell.border, score_cell.alignment = border, center_align
            current_row += 1
        current_row += 1

    # --- SECTION 5: TECHNICAL EXAMPLES ---
    examples = agent_data.get('audit_data', {}).get('technical_examples', [])
    if examples:
        ws.cell(row=current_row, column=1, value="TECHNICAL EXAMPLES").font = subheader_font
        current_row += 1
        
        headers = ['#', 'Client', 'PIN', 'Issue Type', 'Customer Issue', 'Agent Action', 'Outcome', 'Assessment', 'Severity']
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=h)
            cell.fill, cell.font, cell.border, cell.alignment = header_fill, header_font, border, center_align
        
        current_row += 1
        for idx, ex in enumerate(examples, 1):
            data = [
                ex.get('example_number', idx), ex.get('client_name', 'N/A'), ex.get('pin_number', 'N/A'),
                ex.get('issue_type', 'N/A'), ex.get('customer_issue', 'N/A'), ex.get('agent_action', 'N/A'),
                ex.get('outcome', 'N/A'), ex.get('assessment', 'N/A'), ex.get('severity', 'N/A')
            ]
            for c_idx, value in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=value)
                cell.border, cell.alignment = border, left_align
                # Severity Coloring
                if c_idx == 9: 
                    sev_colors = {'Critical': "F8D7DA", 'Major': "FFF3CD", 'Moderate': "D1ECF1", 'Minor': "D4EDDA"}
                    if value in sev_colors:
                        cell.fill = PatternFill(start_color=sev_colors[value], end_color=sev_colors[value], fill_type="solid")
            current_row += 1
        current_row += 1

    # --- SECTION 6: STRENGTHS & DEVELOPMENT ---
    for title, key in [("KEY STRENGTHS", "key_strengths"), ("AREAS FOR DEVELOPMENT", "key_development_areas")]:
        ws.cell(row=current_row, column=1, value=title).font = title_font
        current_row += 1
        items = agent_data.get('audit_data', {}).get(key, [])
        for i, item in enumerate(items, 1):
            ws.cell(row=current_row, column=1, value=f"{i}.")
            content_cell = ws.cell(row=current_row, column=2, value=item)
            content_cell.alignment, content_cell.border = left_align, border
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            current_row += 1
        current_row += 1

    # --- SECTION 7: SECURITY & PIN PROTOCOL ---
    ws.cell(row=current_row, column=1, value="SECURITY & PIN PROTOCOL ANALYSIS").font = title_font
    current_row += 1
    pin_feedback = agent_data.get('audit_data', {}).get('pin_protocol_feedback', 'No feedback available')
    ws.cell(row=current_row, column=1, value=pin_feedback).alignment = left_align
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=4)
    current_row += 3

    # --- SECTION 8: OVERALL ASSESSMENT ---
    ws.cell(row=current_row, column=1, value="OVERALL ASSESSMENT").font = title_font
    current_row += 1
    assessment = agent_data.get('audit_data', {}).get('overall_assessment', 'No assessment available')
    ws.cell(row=current_row, column=1, value=assessment).alignment = left_align
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+4, end_column=4)
    current_row += 6

    # --- SECTION 9: CRITICAL INCIDENTS (Conditional) ---
    critical = agent_data.get('audit_data', {}).get('critical_incidents', [])
    if critical:
        ws.cell(row=current_row, column=1, value="CRITICAL INCIDENTS").font = Font(bold=True, size=16, color="DC3545")
        current_row += 1
        for i, incident in enumerate(critical, 1):
            ws.cell(row=current_row, column=1, value=f"{i}.")
            cell = ws.cell(row=current_row, column=2, value=incident)
            cell.alignment, cell.border = left_align, border
            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            current_row += 1

    # Final adjustments
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    
    wb.save(output_path)
    return output_path

# --- UI DISPLAY ---
def display_results(audit_data):
    """Display audit results in the Streamlit UI"""
    
    # Overall Score
    score = min(float(audit_data.get('overall_score', 0)), 10.0)
    
    # Chat Volume Metrics
    
        st.markdown("### 📊 Chat Volume Performance")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
        with col2:
        with col3:
            st.metric("Performance", f"{perf_pct}%")
        with col4:
            rating_emoji = {
                'Excellent': '🟢',
                'Good': '🟡',
                'Needs Improvement': '🟠',
                'Below Standard': '🔴'
            }.get(rating, '⚪')
            st.metric("Rating", f"{rating_emoji} {rating}")
        
        st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown(f"""
        <div style='text-align: center; padding: 2rem; background-color: #e8f4f8; border-radius: 10px; border: 3px solid #1f77b4;'>
            <h1 style='color: #1f77b4; margin: 0;'>{score}/10.0</h1>
            <p style='font-size: 1.2rem; color: #666; margin: 0.5rem 0 0 0;'>Overall Performance Score</p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Overall Assessment
    st.markdown("### 📋 Overall Assessment")
    st.info(audit_data.get("overall_assessment", "No assessment available"))
    
    # Metrics
    st.markdown("### 📊 Performance Metrics (Out of 5.0)")
    metrics = audit_data.get("metrics", {})
    cols = st.columns(len(metrics))
    
    for i, (key, value) in enumerate(metrics.items()):
        display_val = min(float(value), 5.0)
        metric_name = key.replace("_", " ").title()
        
        # Color coding based on score
        if display_val >= 4.0:
            color = "🟢"
        elif display_val >= 3.0:
            color = "🟡"
        else:
            color = "🔴"
        
        cols[i].metric(f"{color} {metric_name}", f"{display_val}/5.0")
    
    st.markdown("---")
    
    # PIN Protocol Feedback
    st.markdown("### 🔐 Security & PIN Verification Analysis")
    st.warning(audit_data.get("pin_protocol_feedback", "No feedback available"))
    
    # Strengths and Development Areas
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### ✅ Key Strengths")
        strengths = audit_data.get("key_strengths", [])
        if strengths:
            for strength in strengths:
                st.markdown(f"""
                <div class='strength-item'>
                    <strong>✓</strong> {strength}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No strengths identified")
    
    with col2:
        st.markdown("### 🎯 Areas for Development")
        dev_areas = audit_data.get("key_development_areas", [])
        if dev_areas:
            for area in dev_areas:
                st.markdown(f"""
                <div class='improvement-item'>
                    <strong>→</strong> {area}
                </div>
                """, unsafe_allow_html=True)
        else:
            st.info("No development areas identified")
    
    st.markdown("---")
    
    # Critical Incidents (if any) - show prominently
    critical = audit_data.get("critical_incidents", [])
    if critical:
        st.markdown("###⚠️ Critical Incidents")
        for incident in critical:
            st.markdown(f"""
            <div class='critical-item'>
                <strong>⚠️</strong> {incident}
            </div>
            """, unsafe_allow_html=True)
        st.markdown("---")
    
    # Technical Examples
    st.markdown("### 🔧 Detailed Technical Examples")
    examples = audit_data.get("technical_examples", [])
    
    if examples:
        # Group by severity
        severity_filter = st.selectbox(
            "Filter by Severity:",
            ["All", "Critical", "Major", "Moderate", "Minor"]
        )
        
        filtered_examples = examples if severity_filter == "All" else [
            ex for ex in examples if ex.get('severity', '') == severity_filter
        ]
        
        for i, example in enumerate(filtered_examples, 1):
            severity = example.get('severity', 'N/A')
            severity_emoji = {
                'Critical': '🔴',
                'Major': '🟠',
                'Moderate': '🟡',
                'Minor': '🟢'
            }.get(severity, '⚪')
            
            with st.expander(f"{severity_emoji} Example {i}: {example.get('issue_type', 'N/A')} - {example.get('customer_issue', 'N/A')[:50]}..."):
                st.markdown(f"**Client Name:** {example.get('client_name', 'N/A')}")
                st.markdown(f"**PIN Number:** {example.get('pin_number', 'N/A')}")
                st.markdown(f"**Customer Issue:** {example.get('customer_issue', 'N/A')}")
                st.markdown(f"**Agent Action:** {example.get('agent_action', 'N/A')}")
                st.markdown(f"**PIN Handled Well:** {example.get('pin_handled_well', 'N/A')}")
                st.markdown(f"**Outcome:** {example.get('outcome', 'N/A')}")
                st.markdown(f"**Assessment:** {example.get('assessment', 'N/A')}")
                st.markdown(f"**Improvement Recommendation:** {example.get('improvement', 'N/A')}")
                st.markdown(f"**Severity:** {severity}")
    else:
        st.info("No technical examples available")
    
    # Performance Trends
    trends = audit_data.get("performance_trends", {})
    if trends:
        st.markdown("---")
        st.markdown("### 📈 Performance Trends")
        for key, value in trends.items():
            trend_name = key.replace('_', ' ').title()
            st.markdown(f"**{trend_name}:** {value}")
    
    # Recommended Training
    training = audit_data.get("recommended_training", [])
    if training:
        st.markdown("---")
        st.markdown("### 🎓 Recommended Training")
        for i, recommendation in enumerate(training, 1):
            st.markdown(f"{i}. {recommendation}")
    
    # Standout Moments
    standout = audit_data.get("standout_moments", [])
    if standout:
        st.markdown("---")
        st.markdown("### ⭐ Standout Moments")
        for moment in standout:
            st.success(moment)

# --- MAIN APP FLOW ---
def main():
    # Header
    st.markdown("<h1 class='main-header'>🤖 HostAfrica AI Auditor - Enhanced Edition V2</h1>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #666;'>Comprehensive Performance Analysis with AI-Powered Insights & Chat Volume Metrics</p>", unsafe_allow_html=True)
    
    # Sidebar
    st.sidebar.title("🏢 Agent Management")
    st.sidebar.markdown("---")
    
    # Configuration settings
    with st.sidebar.expander("⚙️ Review Settings"):
        review_period = st.number_input(
            "Review Period (days)", 
            min_value=30, 
            max_value=180, 
            value=REVIEW_PERIOD_DAYS,
            help="Default: 90 days (quarterly review)"
        )
        expected_weekly = st.number_input(
            "Expected Chats per Week", 
            min_value=20, 
            max_value=100, 
            value=EXPECTED_CHATS_PER_5_DAYS,
            help="Based on 5-day work week"
        )
    
    st.sidebar.markdown("---")
    
    # Add new agent
    st.sidebar.subheader("Add New Agent")
    new_agent_name = st.sidebar.text_input("Agent Name", key="new_agent_input")
    
    if st.sidebar.button("➕ Add Agent", use_container_width=True):
        if new_agent_name:
            if new_agent_name not in st.session_state.agents:
                agent = get_initial_agent(new_agent_name)
                agent['review_period_days'] = review_period
                st.session_state.agents[new_agent_name] = agent
                st.sidebar.success(f"Added {new_agent_name}")
                st.rerun()
            else:
                st.sidebar.warning("Agent already exists")
        else:
            st.sidebar.error("Please enter an agent name")
    
    st.sidebar.markdown("---")
    
    # Agent list
    agent_list = list(st.session_state.agents.keys())
    
    if not agent_list:
        st.info("👈 Please add an agent using the sidebar to begin")
        
        # Show performance standards
        st.markdown("### 📊 Performance Standards")
        st.info(f"""
        **Quarterly Review Standards (90 days):**
        - Expected chats per 5-day week: **{EXPECTED_CHATS_PER_5_DAYS}**
        - Expected chats for 90-day period: **~{EXPECTED_CHATS_QUARTERLY}**
        - Rating scale:
          - 🟢 Excellent: 100%+ of expected
          - 🟡 Good: 80-99% of expected
          - 🟠 Needs Improvement: 60-79% of expected
          - 🔴 Below Standard: <60% of expected
        """)
        return
    
    # Select agent
    st.sidebar.subheader("Select Agent")
    selected_agent = st.sidebar.selectbox(
        "Choose an agent:",
        agent_list,
        key="agent_selector"
    )
    
    # Remove agent button
    if st.sidebar.button(f"🗑️ Remove {selected_agent}", use_container_width=True):
        del st.session_state.agents[selected_agent]
        st.rerun()
    
    st.sidebar.markdown("---")
    st.sidebar.info(f"**Total Agents:** {len(agent_list)}")
    
    # Main content area
    agent = st.session_state.agents[selected_agent]
    
    # Tabs
    tab1, tab2 = st.tabs(["🤖 Audit Interface", "📈 Detailed Results & Export"])
    
    with tab1:
        st.subheader(f"Analyzing: {selected_agent}")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            zip_file = st.file_uploader(
                "Upload tawk.to ZIP file containing chat transcripts",
                type="zip",
                help="Upload the exported ZIP file from tawk.to"
            )
        
        with col2:
            if agent.get("audit_timestamp"):
                st.metric("Last Audit", agent["audit_timestamp"].strftime("%Y-%m-%d %H:%M"))
            if agent.get("total_chats"):
                st.metric("Chats Analyzed", agent["total_chats"])
        
        if zip_file:
            st.success(f"✅ File uploaded: {zip_file.name}")
            
            if st.button("🚀 Run Comprehensive Audit", use_container_width=True, type="primary"):
                with st.spinner(f"🔍 AI is analyzing transcripts for {selected_agent}..."):
                    
                    # Progress bar
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Extract transcripts
                    status_text.text("📂 Extracting transcripts from ZIP file...")
                    progress_bar.progress(20)
                    transcripts, metadata = get_agent_transcripts(zip_file, selected_agent)
                    
                    if not transcripts:
                        st.error(f"❌ No chats found for agent '{selected_agent}' in the uploaded file.")
                        st.info("Please verify that the agent name matches exactly as it appears in the chat transcripts.")
                        return
                    
                    # Update progress
                    status_text.text(f"✅ Found {len(transcripts)} chats for {selected_agent}")
                    progress_bar.progress(40)
                    
                    # Run AI audit
                    status_text.text("🤖 Running AI-powered comprehensive analysis...")
                    progress_bar.progress(60)
                    
                    audit_result = run_comprehensive_audit(
                        transcripts, 
                        selected_agent, 
                        len(transcripts),
                        agent.get('review_period_days', REVIEW_PERIOD_DAYS)
                    )
                    
                    if audit_result:
                        # Update agent data
                        agent["audit_data"] = audit_result
                        agent["total_chats"] = len(transcripts)
                        agent["audit_timestamp"] = datetime.now()
                        agent["raw_transcripts"] = transcripts[:10]  # Store first 10 for reference
                        
                        progress_bar.progress(100)
                        status_text.text("✅ Analysis complete!")
                        
                        st.success(f"🎉 Successfully analyzed {len(transcripts)} chats!")
                        st.balloons()
                        
                        # Show quick summary
                        st.markdown("### Quick Summary")
                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric("Overall Score", f"{audit_result.get('overall_score', 0)}/10")
                        col2.metric("Strengths", len(audit_result.get('key_strengths', [])))
                        col3.metric("Dev Areas", len(audit_result.get('key_development_areas', [])))
                        
                        
                        st.info("👉 Switch to the 'Detailed Results & Export' tab to view the full report and download PDF")
                    else:
                        st.error("❌ Analysis failed. Please try again or check the error messages above.")
        else:
            st.info("📤 Please upload a ZIP file containing tawk.to chat transcripts to begin analysis")
            
            # Show expected standards
            st.markdown("### 📊 Performance Standards")
            st.info(f"""
            For a **90-day quarterly review**, we expect:
            - **~{EXPECTED_CHATS_QUARTERLY} total chats** (based on {EXPECTED_CHATS_PER_5_DAYS} chats per 5-day week)
            - Consistent daily performance across the review period
            - Quality over quantity, but volume is also measured
            """)
    
    with tab2:
        if agent.get("audit_data"):
            # Display results
            display_results(agent["audit_data"])
            
            st.markdown("---")
            
            # Export section
            st.markdown("### 📥 Export Report")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Download PDF
                if st.button("📄 Generate PDF Report", use_container_width=True, type="primary"):
                    with st.spinner("Generating PDF report..."):
                        # Generate PDF
                        pdf_path = f"/tmp/performance_review_{selected_agent}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                        generate_pdf_report(agent, selected_agent, pdf_path)
                        
                        # Read PDF
                        with open(pdf_path, "rb") as f:
                            pdf_data = f.read()
                        
                        # Download button
                        st.download_button(
                            label="⬇️ Download PDF",
                            data=pdf_data,
                            file_name=f"HostAfrica_Performance_Review_{selected_agent}_{datetime.now().strftime('%Y%m%d')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                        
                        st.success("✅ PDF report generated!")
            
            with col2:
                # Download Excel
                if st.button("📊 Generate Excel Report", use_container_width=True, type="primary"):
                    with st.spinner("Generating Excel report..."):
                        # Generate Excel
                        excel_path = f"/tmp/performance_review_{selected_agent}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        generate_excel_report(agent, selected_agent, excel_path)
                        
                        # Read Excel
                        with open(excel_path, "rb") as f:
                            excel_data = f.read()
                        
                        # Download button
                        st.download_button(
                            label="⬇️ Download Excel",
                            data=excel_data,
                            file_name=f"HostAfrica_Performance_Review_{selected_agent}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        st.success("✅ Excel report generated!")
            
            with col3:
                # Download JSON
                json_data = json.dumps(agent["audit_data"], indent=2)
                st.download_button(
                    label="📋 Download JSON",
                    data=json_data,
                    file_name=f"audit_data_{selected_agent}_{datetime.now().strftime('%Y%m%d')}.json",
                    mime="application/json",
                    use_container_width=True
                )
        else:
            st.info("ℹ️ No audit data available. Please run an audit in the 'Audit Interface' tab first.")

if __name__ == "__main__":
    main()
